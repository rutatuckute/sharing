import requests, json
import pandas as pd
import openpyxl
from openpyxl import *
import time
import datetime
from pandas.io.json import json_normalize

current_time = time.time()
daily_time = datetime.datetime.fromtimestamp(current_time).strftime('%Y%m%d_%H%M%S')

url = "https://skyscanner-skyscanner-flight-search-v1.p.rapidapi.com/apiservices/browsequotes/v1.0/FR/EUR/en-US/PARI-sky/VILN-sky/2020-12-20/2020-12-27"

headers = {
    'x-rapidapi-key': "your_API_key",
    'x-rapidapi-host': "skyscanner-skyscanner-flight-search-v1.p.rapidapi.com"
    }

response = requests.request("GET", url, headers=headers)
take_me_home = response.json()
quotes = json_normalize(take_me_home['Quotes'])


file_path = "C:\\Users\\ruttuc\\Desktop\\take_me_home\\"
file_save = "/quotes_hoard.xlsx"

quotes['QueryDate'] = daily_time

max_row = load_workbook(file_path+file_save)['data'].max_row

book = load_workbook(file_path+file_save)
writer = pd.ExcelWriter(file_path+file_save, engine='openpyxl') 
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

quotes.to_excel(writer, 'data', startrow=max_row, index=False, header=False)
writer.save()

